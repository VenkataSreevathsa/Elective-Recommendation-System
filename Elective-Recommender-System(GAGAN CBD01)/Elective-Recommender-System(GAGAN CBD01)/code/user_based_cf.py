import os
import openpyxl

def IsScore(sheet, r, c):
    if sheet.cell(row=r + 1, column=c).value != 'NT' \
            and sheet.cell(row=r + 1, column=c).value != 'I' \
            and sheet.cell(row=r + 1, column=c).value != 'XX' \
            and sheet.cell(row=r + 1, column=c).value != None:
        return True
    else:
        return False

def scr_sub(sheet, u, i):
    return sheet.cell(row=u + 1, column=i).value

print("Collaborative Filtering User-Based Algorithm for Grade Prediction")
wb = openpyxl.load_workbook('C:/Users/Administrator/Desktop/Elective-Recommender-System(GAGAN CBD01)/Elective-Recommender-System(GAGAN CBD01)/code/data/student.xlsx')

sheetname = wb.sheetnames

print("Student's score matrix(roll no in first column)\n")
sheet = wb[sheetname[0]]

for i in range(2, sheet.max_row + 1, 1):
    for j in range(1, sheet.max_column + 1, 1):
        if sheet.cell(row=i, column=j).value != None:
            if j == 1:
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(sheet.cell(row=i, column=j).value, end="  | ")
                else:
                    print(sheet.cell(row=i, column=j).value, end=" |  ")
            else:
                print(sheet.cell(row=i, column=j).value, end="  ")
            if sheet.cell(row=i, column=j).value != 'I' and sheet.cell(row=i, column=j).value != 'XX' and sheet.cell(row=i, column=j).value != 'NT':
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(end=" ")
            if sheet.cell(row=i, column=j).value == 'I':
                print(end=" ")
        else:
            print("-", end="   ")
    print()

studid = int(input("Enter roll no. : "))
coursesList = []
for i in range(2, sheet.max_column + 1):
    if (i - 2) % 9 == 0:
        print()
    print(sheet.cell(row=1, column=i).value, end=", ")
    coursesList.append(sheet.cell(row=1, column=i).value)
print("\n")

# Calculate average grades for a student
def getAvgGrades(sheet, student):
    sum = 0
    cnt = 0
    for i in range(2, sheet.max_column):
        if IsScore(sheet, student, i):
            sum += int(scr_sub(sheet, student, i))
            cnt += 1
    if cnt != 0:
        return round(float(sum / cnt), 3)
    else:
        return 0

# Calculate similarity between two students
def similarity(u, v, sheet, avgU):
    sum1 = 0
    sum2 = 0
    sum3 = 0
    avgV = getAvgGrades(sheet, v)

    for i in range(2, sheet.max_column + 1):
        if IsScore(sheet, u, i) and IsScore(sheet, v, i):
            sum1 += (scr_sub(sheet, u, i) - avgU) * (scr_sub(sheet, v, i) - avgV)
            sum2 += (scr_sub(sheet, u, i) - avgU) * (scr_sub(sheet, u, i) - avgU)
            sum3 += (scr_sub(sheet, v, i) - avgV) * (scr_sub(sheet, v, i) - avgV)
    if sum2 != 0 and sum3 != 0:
        return round(float(sum1 / ((sum2 * sum3) ** (1 / 2))), 3)
    else:
        return 0

similar_user_val = []
simi_user = []

avgU = getAvgGrades(sheet, studid)
for i in range(1, sheet.max_row):
    if i != studid:
        sim = similarity(studid, i, sheet, avgU)
        similar_user_val.append([studid, i, sim])
        if sim > 0:
            simi_user.append([studid, i, sim])

# Predicting grades for all courses
predicted_grades = {}
for course in range(2, sheet.max_column + 1):
    sum1 = 0
    sum2 = 0
    for i in range(0, len(simi_user)):
        if IsScore(sheet, int(simi_user[i][1]), course):
            avg = getAvgGrades(sheet, simi_user[i][1])
            sum1 += simi_user[i][2] * (scr_sub(sheet, simi_user[i][1], course) - avg)
            sum2 += simi_user[i][2]
    if sum2 != 0:
        predicted_grade = round(float(sum1 / sum2), 3) + avgU
        predicted_grades[sheet.cell(row=1, column=course).value] = predicted_grade

# Displaying predicted grades for all courses
print("\nPredicted Grades for Student " + str(studid) + ":")
for course_name, grade in predicted_grades.items():
    print(f"{course_name}: {grade}")
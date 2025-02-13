import os
import openpyxl

def IsScore(sheet, r, c):
    if sheet.cell(row=r + 1, column=c).value != 'NT' \
            and sheet.cell(row=r + 1, column=c).value != 'I' \
            and sheet.cell(row=r + 1, column=c).value != 'XX' \
            and sheet.cell(row=r + 1, column=c).value != None:
        return True
    else:
        return False

def scr_sub(sheet, u, i):
    return sheet.cell(row=u + 1, column=i).value

# Placeholder for the similarity function
def similarity(course1, course2, sheet):
    # Implement your similarity calculation logic here
    return 1  # Placeholder value, replace with actual similarity calculation

print("Collaborative Filtering Item-Based Algorithm for Grade Prediction")
wb = openpyxl.load_workbook('C:/Users/Administrator/Downloads/Elective-Recommender-System(GAGAN CBD01)/Elective-Recommender-System(GAGAN CBD01)/Elective-Recommender-System(GAGAN CBD01)/code/data/student.xlsx')
sheetname = wb.sheetnames
print("Student's score matrix (roll no in first column)\n")

sheet = wb[sheetname[1]]
# Print the score matrix
for i in range(2, sheet.max_row + 1, 1):
    for j in range(1, sheet.max_column + 1, 1):
        if sheet.cell(row=i, column=j).value != None:
            if j == 1:
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(sheet.cell(row=i, column=j).value, end="  | ")
                else:
                    print(sheet.cell(row=i, column=j).value, end=" |  ")
            else:
                print(sheet.cell(row=i, column=j).value, end="  ")
            if sheet.cell(row=i, column=j).value != 'I' and sheet.cell(row=i, column=j).value != 'XX' and sheet.cell(row=i, column=j).value != 'NT':
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(end=" ")
            if sheet.cell(row=i, column=j).value == 'I':
                print(end=" ")
        else:
            print("-", end="   ")
    print()

studid = int(input("Enter roll no. : "))
coursesList = []
for i in range(2, sheet.max_column + 1):
    if (i - 2) % 9 == 0:
        print()
    print(sheet.cell(row=1, column=i).value, end=", ")
    coursesList.append(sheet.cell(row=1, column=i).value)
print("\n")

# Predicting grades for all courses
predicted_grades = {}

for course in range(2, sheet.max_column + 1):
    sum1 = 0
    sum2 = 0
    for i in range(2, sheet.max_column + 1):
        if i != course:
            sim = similarity(i, course, sheet)
            if sim > 0 and IsScore(sheet, studid, i):
                sum1 += (sim * scr_sub(sheet, studid, i))
                sum2 += sim
    if sum2 != 0:
        predicted_grades[sheet.cell(row=1, column=course).value] = round(float(sum1 / sum2), 3)
    else:
        predicted_grades[sheet.cell(row=1, column=course).value] = "No prediction available"

# Output predicted grades for all courses
print("\nPredicted Grades for student " + str(studid) + ":")
for course, grade in predicted_grades.items():
    print(f"{course}: {grade}")